import openpyxl
from datetime import datetime


path = r'C:\Users\i506998\OneDrive - SAP SE\Documents\private\Master Degree\Final Project\My research\TD_Ameritrade\data.xlsx'
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

results = open("results.csv", "a")
results.truncate(0)

header = 'MACD(),MACDdiff,MACD().Avg,RSI(),"ExpAverage(close, length = 9)","ExpAverage(close, length = 21)",' \
         '"ExpAverage(close, length = 34)","ExpAverage(close, length = 55)","ExpAverage(close, length = 88)",' \
         '"ExpAverage(close, length = 100)",BollingerBands().UpperBand,BollingerBands().LowerBand,CCI(),' \
         'StochasticFull().FullD,StochasticFull().FullK,imp_volatility,volume,close,GetTime(),result_label'
results.write(header + "\n")

for row in range(sheet_obj.max_row-1):
    if (row + 1) % 2 == 0:
        if sheet_obj.cell(row = row + 1, column = 3).value.__contains__("Close"):
            print('FAIL at ' + str(row))

        cell_obj = sheet_obj.cell(row + 1, column=2)
        raw_data_list = cell_obj.value.split(" ")
        print(raw_data_list)
        raw_data_list = raw_data_list[1:]

        # formatting timestamp
        raw_data_list[raw_data_list.__len__()-1] = raw_data_list[raw_data_list.__len__()-1].split(')')[0]
        time = datetime.utcfromtimestamp(int(raw_data_list[raw_data_list.__len__()-1].replace(',', ''))/1000)
        formated_time = time.hour + time.minute / 60

        raw_data_list[raw_data_list.__len__()-1] = str(round(formated_time, 2))
        raw_data_list = list(map(lambda x: x.replace('N/A', '0'), raw_data_list))
        raw_data_list = list(map(lambda x: x.replace(',', ''), raw_data_list))
        print(raw_data_list)
        s = ','.join(raw_data_list)

        profit_loss = sheet_obj.cell(row + 2, column=7).value
        is_loss = profit_loss.__contains__("(")
        profit_loss = profit_loss.replace("(", "")
        profit_loss = profit_loss.replace(")", "")
        profit_loss = profit_loss.replace("$", "")
        profit_loss = profit_loss.replace(",", "")

        pl_size = float(profit_loss)/300
        if pl_size >= 0.1 :
            if is_loss :
                s = s + "," + str('0')
            else:
                s = s + "," + str('1')
        # else:
        #     s = s + "," + str('0')

        results.write(s + "\n")
        # print(s)


