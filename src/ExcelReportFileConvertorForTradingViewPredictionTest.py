import openpyxl
import pandas as pd
from sklearn.preprocessing import RobustScaler, StandardScaler
from sklearn.svm import SVC
from datetime import datetime
from io import StringIO
import pickle


FINALIZED_MODEL_SAV = 'finalized_model_tradingview_2000_12_000_percent.sav'
path = r'C:\Users\i506998\OneDrive - SAP SE\Documents\private\Master Degree\Final Project\My research\TradingView\ERIK_TEST_MACD_Strategy_List_of_Trades_2022-08-21_deep_data_from_excel.xlsx'
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

results = open("results_new_tradingview_025_percents_predict_result_new.csv", "a")
results.truncate(0)

header = 'MACD(),MACDdiff,MACD().Avg,RSI(),"ExpAverage(close, length = 9)","ExpAverage(close, length = 21)",' \
         '"ExpAverage(close, length = 34)","ExpAverage(close, length = 55)","ExpAverage(close, length = 88)",' \
         '"ExpAverage(close, length = 100)",BollingerBands().UpperBand,BollingerBands().LowerBand,CCI(),' \
         'StochasticFull().FullD,StochasticFull().FullK,imp_volatility,volume,close,GetTime(),predicted_result,result_label,profit_loss'
results.write(header + "\n")

print(str(sheet_obj.max_row))

clf_svm: SVC = pickle.load(open(FINALIZED_MODEL_SAV, 'rb'))
standardScaler: StandardScaler = pickle.load(open("standardScaler.sav", 'rb'))

counter_of_0 = 0
counter_of_1 = 0

counter_of_1p_1r = 0
counter_of_1p_0r = 0
counter_of_0p_1r = 0
counter_of_0p_0r = 0

for row in range(2, sheet_obj.max_row):
    if (row) % 2 == 1:
        print('row number ' + str(row))
        if sheet_obj.cell(row = row, column = 2).value.__contains__("Exit"):
            print('FAIL at ' + str(row))

        cell_obj = sheet_obj.cell(row, column=3)
        raw_data_list = cell_obj.value.split(" ")
        # print(raw_data_list)
        raw_data_list = raw_data_list[1:]

        # formatting timestamp
        raw_data_list[raw_data_list.__len__()-1] = raw_data_list[raw_data_list.__len__()-1].split(')')[0]
        time = datetime.utcfromtimestamp(int(raw_data_list[raw_data_list.__len__()-1].replace(',', ''))/1000)
        formated_time = time.hour + time.minute / 60

        raw_data_list[raw_data_list.__len__()-1] = str(round(formated_time, 2))
        raw_data_list = list(map(lambda x: x.replace('NaN', '0'), raw_data_list))
        raw_data_list = list(map(lambda x: x.replace(',', ''), raw_data_list))
        s = ','.join(raw_data_list)

        header = 'MACD(),MACDdiff,MACD().Avg,RSI(),"ExpAverage(close, length = 9)","ExpAverage(close, length = 21)",' \
         '"ExpAverage(close, length = 34)","ExpAverage(close, length = 55)","ExpAverage(close, length = 88)",' \
         '"ExpAverage(close, length = 100)",BollingerBands().UpperBand,BollingerBands().LowerBand,CCI(),' \
         'StochasticFull().FullD,StochasticFull().FullK,imp_volatility,volume,close,GetTime()'

        alertBodyFormattedWithHeader = header + '\n' + s

        array = pd.read_csv(StringIO(alertBodyFormattedWithHeader))

        df_scaled = standardScaler.transform(array)
        # print('df_scaled:' + str(df_scaled))

        reshaped_scaled_row = df_scaled[0].reshape(1, -1)
        # print('reshaped_scaled_row: ' + str(reshaped_scaled_row))

        prediction = clf_svm.predict(reshaped_scaled_row)

        if prediction[0] == 0:
            # print(prediction[0])
            s = s + "," + str('0')
            counter_of_0 += 1
        else:
            # print(prediction[0])
            s = s + "," + str('1')
            counter_of_1 += 1

        profit_loss = sheet_obj.cell(row, column=7).value
        is_loss = True if profit_loss<0 else False
        # profit_loss = profit_loss.replace("(", "").replace(")", "").replace("$", "").replace(",", "")

        pl_size = abs(float(profit_loss))/sheet_obj.cell(row, column=5).value
        if pl_size >= 0.0025 :
            if is_loss :
                s = s + "," + str('0')
                if( prediction[0]==1 ) :
                    counter_of_1p_0r += 1;
                else:
                    counter_of_0p_0r += 1;
            else:
                s = s + "," + str('1')
                if( prediction[0] == 1 ) :
                    counter_of_1p_1r += 1;
                else:
                    counter_of_0p_1r += 1;
        else:
            s = s + "," + str('0')
            if( prediction[0] == 1 ) :
                counter_of_1p_0r += 1;
            else:
                counter_of_0p_0r += 1;

        s = s + "," + str(profit_loss) + "\n"

        results.write(s)


print('counter_of_0 ' + str(counter_of_0))
print('counter_of_1 ' + str(counter_of_1))
print('counter_of_1p_1r ' + str(counter_of_1p_1r))
print('counter_of_1p_0r ' + str(counter_of_1p_0r))
print('counter_of_0p_1r ' + str(counter_of_0p_1r))
print('counter_of_0p_0r ' + str(counter_of_0p_0r))




