import openpyxl
from datetime import datetime


path = r'C:\Users\i506998\OneDrive - SAP SE\Documents\private\Master Degree\Final Project\My research\TradingView\ERIK_TEST_MACD_Strategy_List_of_Trades_2022-08-21_deep_data_from_excel.xlsx'
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

results = open("results_new_tradingview_000_percents_new.csv", "a")
results.truncate(0)

header = 'MACD(),MACDdiff,MACD().Avg,RSI(),"ExpAverage(close, length = 9)","ExpAverage(close, length = 21)",' \
         '"ExpAverage(close, length = 34)","ExpAverage(close, length = 55)","ExpAverage(close, length = 88)",' \
         '"ExpAverage(close, length = 100)",BollingerBands().UpperBand,BollingerBands().LowerBand,CCI(),' \
         'StochasticFull().FullD,StochasticFull().FullK,imp_volatility,volume,close,GetTime(),result_label,profit_loss'
results.write(header + "\n")

print(str(sheet_obj.max_row))

for row in range(2, sheet_obj.max_row):
    if (row) % 2 == 1:
        print('row number ' + str(row))
        if sheet_obj.cell(row = row, column = 2).value.__contains__("Exit"):
            print('FAIL at ' + str(row))

        cell_obj = sheet_obj.cell(row, column=3)
        raw_data_list = cell_obj.value.split(" ")
        # print(raw_data_list)
        raw_data_list = raw_data_list[1:]

        # formatting timestamp
        raw_data_list[raw_data_list.__len__()-1] = raw_data_list[raw_data_list.__len__()-1].split(')')[0]
        time = datetime.utcfromtimestamp(int(raw_data_list[raw_data_list.__len__()-1].replace(',', ''))/1000)
        formated_time = time.hour + time.minute / 60

        raw_data_list[raw_data_list.__len__()-1] = str(round(formated_time, 2))
        raw_data_list = list(map(lambda x: x.replace('NaN', '0'), raw_data_list))
        raw_data_list = list(map(lambda x: x.replace(',', ''), raw_data_list))
        # print(raw_data_list)
        s = ','.join(raw_data_list)

        profit_loss = sheet_obj.cell(row, column=7).value
        is_loss = True if profit_loss<0 else False
        # profit_loss = profit_loss.replace("(", "").replace(")", "").replace("$", "").replace(",", "")

        pl_size = abs(float(profit_loss))/sheet_obj.cell(row, column=5).value
        # if pl_size >= 0.0025 :
        if is_loss :
            s = s + "," + str('0')
        else:
            s = s + "," + str('1')
        # else:
        #     s = s + "," + str('0')

        s = s + "," + str(profit_loss) + "\n"

        results.write(s)
        # print(s)




